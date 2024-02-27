from io import BytesIO
from django.test import TestCase, Client
from django.urls import reverse
from django.utils.http import urlencode

from datetime import date
import openpyxl

from libraries.models import Book, Library, BookStatus, BookGenders
from persons.models import Author

class BaseTestCase(TestCase):
    def setUp(self):
        self.genders = BookGenders.objects.all()
        self.author1 = Author.objects.create(first_name = 'Darwin',
                                             last_name = 'Lantigua',
                                             rnc = '402-3070960-8',
                                             birthday = date(2000, 1, 8),
                                             alias = 'Esnaire'
                                             )
        self.library1 = Library.objects.create(name = 'libreria1', 
                                               location = 'Tenares',
                                               rnc = '123-1234567-1')
        self.library2 = Library.objects.create(name = 'libreria2', 
                                               location = 'Salcedo',
                                               rnc = '123-1234567-2')
        self.status = BookStatus.objects.all()
        self.book1 = Book.objects.create(title = 'Our love is live',
                                         published_date = date(2020, 6, 2),
                                         isbn = '123-1-123-12345-1',
                                         gender = self.genders[7],
                                         quantity = 5,
                                         rent_price = 300,
                                         sale_price = 500,
                                         author = self.author1,
                                         library = self.library1,
                                         status = self.status[0],
                                         )
        self.client = Client()

class IndexViewTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.book2 = Book.objects.create(title = 'Gamer',
                                         published_date = date(2020, 6, 1),
                                         isbn = '123-1-123-12345-2',
                                         gender = self.genders[5],
                                         quantity = 6,
                                         rent_price = 200,
                                         sale_price = 600,
                                         author = self.author1,
                                         library = self.library2,
                                         status = self.status[0],
                                         )
        self.book3 = Book.objects.create(title = 'Touch your heart',
                                         published_date = date(2020, 6, 3),
                                         isbn = '123-1-123-12345-3',
                                         gender = self.genders[7],
                                         quantity = 6,
                                         rent_price = 250,
                                         sale_price = 450,
                                         author = self.author1,
                                         library = self.library1,
                                         status = self.status[1],
                                         )
        self.book4 = Book.objects.create(title = 'Learn to code',
                                         published_date = date(2020, 6, 4),
                                         isbn = '123-1-123-12345-4',
                                         gender = self.genders[1],
                                         quantity = 6,
                                         rent_price = 100,
                                         sale_price = 300,
                                         author = self.author1,
                                         library = self.library2,
                                         status = self.status[11],
                                         )
        self.url = reverse('libraries:book_index')

    def test_get_index(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'book/book_index.html')
        self.assertEqual(len(response.context['model_list']), 4)

    def test_filter_by_value1(self):
        params = {'filter_value': 'our l', 'filter_gender': 0, 'filter_status': 0}
        query_string = urlencode(params)
        url = f"{reverse('libraries:book_index')}?{query_string}"

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['model_list']), 1)
        self.assertEqual(response.context['model_list'][0], self.book1)

    def test_filter_by_value2(self):
        params = {'filter_value': 'libreria2', 'filter_gender': 0, 'filter_status': 0}
        query_string = urlencode(params)
        url = f"{reverse('libraries:book_index')}?{query_string}"

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['model_list']), 2)
        self.assertEqual(response.context['model_list'][0], self.book2)
        self.assertEqual(response.context['model_list'][1], self.book4)

    def test_filter_by_gender(self):
        params = {'filter_value': '', 'filter_gender': 8, 'filter_status': 0}
        query_string = urlencode(params)
        url = f"{reverse('libraries:book_index')}?{query_string}"

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['model_list']), 2)
        self.assertEqual(response.context['model_list'][0], self.book1)
        self.assertEqual(response.context['model_list'][1], self.book3)

    def test_filter_by_status(self):
        params = {'filter_value': '', 'filter_gender': 0, 'filter_status': 1}
        query_string = urlencode(params)
        url = f"{reverse('libraries:book_index')}?{query_string}"

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['model_list']), 2)
        self.assertEqual(response.context['model_list'][0], self.book1)
        self.assertEqual(response.context['model_list'][1], self.book2)

def create_book(title, published_date, isbn, gender, quantity, rent_price, 
                sale_price, author, library, status):
    return {
        'title': title,
        'published_date': published_date,
        'isbn': isbn,
        'gender': gender,
        'quantity': quantity,
        'rent_price': rent_price,
        'sale_price': sale_price,
        'author': author,
        'library': library,
        'status': status,
    }

class CreateViewTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.url = reverse('libraries:book_create')

    def test_get_view(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'book/book_create_form.html')

    def test_valid_data_post(self):
        data = create_book('Red String', date(2020, 6, 2), 
                           '123-1-123-54321-1', self.genders[0].pk, 5, 300, 
                           500, self.author1.pk, self.library1.pk, 
                           self.status[0].pk)
        response = self.client.post(self.url, data, follow=True)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(Book.objects.count(), 2)
        book_created = Book.objects.get(title = 'Red String')
        self.assertEqual(book_created.author, self.author1)
        self.assertEqual(book_created.library, self.library1)
        self.assertEqual(book_created.isbn, '1231123543211')

    def test_duplicate_isbn(self):
        data = create_book('Alpha', date(2018, 8, 6), '123-1-12312345-1',
                           self.genders[0].pk, 7, 200, 600, self.author1.pk, 
                           self.library1.pk, self.status[1].pk)
        response = self.client.post(self.url, data, follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'book/book_create_form.html')
        self.assertEqual(Book.objects.count(), 1)
        self.assertFalse(Book.objects.filter(title='Alpha').exists())

    def test_invalid_isbn(self):
        data = create_book('Red String', date(2020, 6, 2), 
                           '123-1-12-12345-1', self.genders[0].pk, 5, 300, 
                           500, self.author1.pk, self.library1.pk, 
                           self.status[0].pk)
        response = self.client.post(self.url, data, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Book.objects.count(), 1)
        self.assertFalse(Book.objects.filter(title='Red String')
                         .exists())

class UpdateViewTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.book2 = Book.objects.create(title = 'Alpha',
                                         published_date = date(2020, 6, 2),
                                         isbn = '111-2-333-44444-5',
                                         gender = self.genders[0],
                                         quantity = 7,
                                         rent_price = 200,
                                         sale_price = 600,
                                         author = self.author1,
                                         library = self.library1,
                                         status = self.status[0]
                                         )
        self.url = reverse('libraries:book_update', args=[self.book2.pk])

    def test_get_view(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'book/book_update.html')

    def test_update_to_duplicate_isbn(self):
        # Submit the form with updated data
        updated_data = create_book(self.book2.title, 
                                   self.book2.published_date, self.book1.isbn,
                                   self.book2.gender, self.book2.quantity, 
                                   self.book2.rent_price, 
                                   self.book2.sale_price, 
                                   self.book2.author.pk, 
                                   self.book2.library.pk, 
                                   self.book2.status.pk)
        response = self.client.post(self.url, updated_data, follow=True)
        self.assertEqual(response.status_code, 200)

        # Refresh the person from the database
        self.book2.refresh_from_db()
        self.assertEqual(self.book2.isbn, '1112333444445')

    def test_update_all_fields_except_isbn(self):
        # Submit the form with updated data
        updated_data = create_book('Blood bond', date(2017, 3, 3), 
                                   self.book2.isbn, self.genders[1].pk, 10, 
                                   150, 300, self.book2.author.pk, 
                                   self.book2.library.pk, self.status[1].pk)
        response = self.client.post(self.url, updated_data, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'book/book_index.html')

        # Refresh the person from the database
        self.book2.refresh_from_db()
        self.assertEqual(self.book2.title, 'Blood bond')

class DetailViewTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.url = reverse('libraries:book_detail', args=[self.book1.pk])
    
    def test_get_view(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'book/book_detail.html')
        self.assertEqual(response.context['model'], self.book1)

class ExcelDownloadViewTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.book2 = Book.objects.create(title = 'Gamer',
                                         published_date = date(2020, 6, 1),
                                         isbn = '123-1-123-12345-2',
                                         gender = self.genders[5],
                                         quantity = 6,
                                         rent_price = 200,
                                         sale_price = 600,
                                         author = self.author1,
                                         library = self.library2,
                                         status = self.status[0],
                                         )
        self.book3 = Book.objects.create(title = 'Touch your heart',
                                         published_date = date(2020, 6, 3),
                                         isbn = '123-1-123-12345-3',
                                         gender = self.genders[7],
                                         quantity = 6,
                                         rent_price = 250,
                                         sale_price = 450,
                                         author = self.author1,
                                         library = self.library1,
                                         status = self.status[1],
                                         )
        self.book4 = Book.objects.create(title = 'Learn to code',
                                         published_date = date(2020, 6, 4),
                                         isbn = '123-1-123-12345-4',
                                         gender = self.genders[1],
                                         quantity = 6,
                                         rent_price = 100,
                                         sale_price = 300,
                                         author = self.author1,
                                         library = self.library2,
                                         status = self.status[11],
                                         )
        self.url = reverse('libraries:download_book_excel')

    def test_default_get_view(self):
        params = {'library_id': 0, 'filter_status_id': 0, 'filter_gender_id': 0, 'filter_value': ''}
        query_string = urlencode(params)
        url = f"{self.url}?{query_string}"

        response = self.client.get(url)
        # Check if the content type is correct
        self.assertEqual(response['Content-Type'], 'application/ms-excel')
        # Check if the Content-Disposition header is set correctly
        self.assertIn('attachment; filename="book.xlsx"', response['Content-Disposition'])
        #import ipdb; ipdb.set_trace()
        # Load workbook from response content
        wb = openpyxl.load_workbook(BytesIO(response.content))
        # Assuming the first worksheet is the one containing data
        ws = wb.active
        # Assuming the first row contains headers
        headers = [cell.value for cell in ws[1]]
        # Assuming the headers are 'Title', 'Gender', 'Quantity', 'Author', 'Library', 'Status'
        expected_headers = ['Title', 'Gender', 'Quantity', 'Author', 'Library', 'Status']
        # Assert headers match
        self.assertEqual(headers, expected_headers)

        row1 = [cell.value for cell in ws[2]]
        expected_row1 = ['Our love is live', 'Romance', 5, 'Darwin Lantigua (Esnaire)', 'libreria1', 'Available']
        self.assertEqual(row1, expected_row1)

        row2 = [cell.value for cell in ws[3]]
        expected_row2 = ['Gamer', 'Fantasy', 6, 'Darwin Lantigua (Esnaire)', 'libreria2', 'Available']
        self.assertEqual(row2, expected_row2)

        row3 = [cell.value for cell in ws[4]]
        expected_row3 = ['Touch your heart', 'Romance', 6, 'Darwin Lantigua (Esnaire)', 'libreria1', 'Spent']
        self.assertEqual(row3, expected_row3)

        row4 = [cell.value for cell in ws[5]]
        expected_row4 = ['Learn to code', 'Non-Fiction', 6, 'Darwin Lantigua (Esnaire)', 'libreria2', 'Reference Only']
        self.assertEqual(row4, expected_row4)

        wb.close()

    def test_filter_by_value1(self):
        params = {'library_id': 0, 'filter_status_id': 0, 'filter_gender_id': 0, 'filter_value': 'ga'}
        query_string = urlencode(params)
        url = f"{self.url}?{query_string}"

        response = self.client.get(url)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        row1 = [cell.value for cell in ws[2]]
        expected_row1 = ['Gamer', 'Fantasy', 6, 'Darwin Lantigua (Esnaire)', 'libreria2', 'Available']
        self.assertEqual(row1, expected_row1)

        wb.close()