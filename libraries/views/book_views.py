from django.views import generic, View
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.template.loader import render_to_string
from django.db.models import Q
import openpyxl
from openpyxl.styles import Alignment, Font

from libraries.models import Book, BookGenders, BookStatus
from libraries.forms.book_forms import BookForm
from libraries.forms.library_forms import LibraryForm
from persons.forms.author_forms import AuthorForm

class IndexView(generic.ListView):
    template_name = 'book/book_index.html'
    context_object_name = 'model_list'
    paginate_by = 4

    def get_queryset(self):
        library_id = self.kwargs.get('library_id')
        filter_value = self.request.GET.get('filter_value', '')
        filter_gender_id = int(self.request.GET.get('filter_gender', '0'))
        filter_status_id = int(self.request.GET.get('filter_status', '0'))

        queryset = Book.objects.all()

        if library_id:
            queryset = queryset.filter(library_id=library_id)

        if filter_status_id != 0:
            queryset = queryset.filter(status_id=filter_status_id)

        if filter_gender_id != 0:
            queryset = queryset.filter(gender_id=filter_gender_id)

        queryset = queryset.filter(
            Q(title__icontains=filter_value) |
            Q(author__first_name__icontains=filter_value) |
            Q(author__last_name__icontains=filter_value) |
            Q(author__alias__icontains=filter_value) |
            Q(library__name__icontains=filter_value)
        )

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super(IndexView, self).get_context_data(**kwargs)

        paginator = Paginator(self.object_list, self.paginate_by)
        page = self.request.GET.get('page')

        try:
            model_list = paginator.page(page)
        except PageNotAnInteger:
            model_list = paginator.page(1)
        except EmptyPage:
            model_list = paginator.page(paginator.num_pages)

        context['model_list'] = model_list
        context['filter_value'] = self.request.GET.get('filter_value', '')
        context['filter_gender'] = self.request.GET.get('filter_gender', '0')
        context['filter_status'] = self.request.GET.get('filter_status', '0')
        context['genders'] = BookGenders.objects.all()
        context['statuses'] = BookStatus.objects.all()

        if 'library_id' in self.kwargs:
            context['library_id'] = self.kwargs['library_id']
        return context
    
    def render_to_response(self, context, **response_kwargs):
        if self.request.is_ajax():
            html = render_to_string('book/book_list_ajax.html', context)
            return JsonResponse({'html': html})
        else:
            return super().render_to_response(context, **response_kwargs)
    
class CreateView(generic.CreateView):
    model = Book
    form_class = BookForm
    template_name_suffix = '_create_form'
    template_name = 'book/book_create_form.html'
    success_url = reverse_lazy('libraries:book_index')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['library_form'] = LibraryForm()
        context['author_form'] = AuthorForm()
        return context
    
class UpdateView(generic.UpdateView):
    model = Book
    form_class = BookForm
    template_name = 'book/book_update.html'
    success_url = reverse_lazy('libraries:book_index')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['library_form'] = LibraryForm()
        context['author_form'] = AuthorForm()
        return context
    
class DetailView(generic.DetailView):
    model = Book
    template_name = 'book/book_detail.html'
    context_object_name = 'model'

def create_modal(request):
    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            instance = form.save()
            return JsonResponse({'status': 'success', 'message': 
                                 'Form submitted successfully'})
        else:
            errors_html = render(request, 'book/book_create_form.html', 
                                 {'form': form}).content.decode('utf-8')
            return JsonResponse({'status': 'error', 'message': 
                                 'Form submission failed', 'errors': 
                                 errors_html}, status=400)
    else:
        form = BookForm()
    
    return render(request, 'book/book_create_form.html', {'form': form})

class ExcelDownloadView(View):
    def get(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="book.xlsx"'

        # Get filter parameters from URL query parameters
        library_id = request.GET.get('library_id')
        filter_status_id = request.GET.get('filter_status_id')
        filter_gender_id = request.GET.get('filter_gender_id')
        filter_value = request.GET.get('filter_value')

        # Initialize queryset with all books
        queryset = Book.objects.all()

        # Apply filters
        if library_id and library_id != '0':
            queryset = queryset.filter(library_id=library_id)

        if filter_status_id and filter_status_id != '0':
            queryset = queryset.filter(status_id=filter_status_id)

        if filter_gender_id and filter_gender_id != '0':
            queryset = queryset.filter(gender_id=filter_gender_id)

        if filter_value:
            queryset = queryset.filter(
                Q(title__icontains=filter_value) |
                Q(author__first_name__icontains=filter_value) |
                Q(author__last_name__icontains=filter_value) |
                Q(author__alias__icontains=filter_value) |
                Q(library__name__icontains=filter_value)
            )

        # Create Excel workbook and add data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Title', 'Gender', 'Quantity', 'Author', 'Library', 'Status'])
        for obj in queryset:
            ws.append([obj.title, obj.gender.__str__() , obj.quantity, 
                       obj.author.__str__(), obj.library.__str__(), 
                       obj.status.__str__()])  # Add data rows

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, 
                                max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', 
                                           vertical='center')
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

        # Set font style of first row to bold
        for cell in ws[1]:  # Iterate over cells in the first row
            cell.font = Font(bold=True)  # Set font style to bold

        wb.save(response)
        return response